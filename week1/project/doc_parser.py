""" 
A program that parses through report files, collecting and
logging data based on the date information contained in filename

COPYRIGHT
    2022(c) Dylan Luttrell
    
LICENSE
    MIT License
"""
__author__ = "Dylan Luttrell"

from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet # import Worksheet for type-hinting
import logging as log
from pathlib import Path
import re

# configure log
FILENAME = "report.log"
LOG_FORMAT = r"%(asctime)s[%(levelname)-6s]:%(message)s"
DATE_FORMAT = "%Y%m%d:%H:%M"
log.basicConfig(level=log.INFO, filename=FILENAME, format=LOG_FORMAT, datefmt=DATE_FORMAT)

# CONSTANTS
DATA_DIR = "data"  # location of data to be parsed, relative to script
WORKING_PATH = Path(__file__).parent # location of script

MONTHS = ("january", "febuary", "march", "april", "may", "june",
          "july", "august", "september", "october", "november", "december")
FN_PATTERN = re.compile(r"^expedia_report_monthly_({})_\d\d\d\d.xlsx?".format( "|".join(MONTHS) ))

def is_valid_filename(filename: str) -> bool:
    """checks whether if valid report file name

    Args:
        filename (str): name of file to be validated

    Returns:
        bool: True if filename is valid, else False
    """
    return bool(re.search(FN_PATTERN, filename))


def get_date_from_filename(filename: str) -> datetime:
    """extracts month and year from filename

    Args:
        filename (str): the filename that date is extracted from

    Returns:
        datetime: month and year data
    """    
    date = re.findall(r"(?<=y_).*?_\d{4}", filename)
    
    if len(date) != 1:
        raise ValueError("invalid filename")
    
    return datetime.strptime(date[0], r"%B_%Y")
    

def log_tab_one(wb: Workbook, date: datetime) -> None:
    """logs data from first tab of report for given date

    Args:
        wb (Workbook): excel data
        date (datetime): date from file name
    """ 
    # NOTE: There is an issue with Pylance and openpyxl. This does not affect execution
    ws: Worksheet = wb[wb.sheetnames[0]]
    
    header = ws["A1:F1"][0]
    
    # iterate through rows containing relevant data
    for row in ws.iter_rows(min_row=2, max_row=15):
        row_id = row[0].value
        
        if isinstance(row_id, str):
            row_id = datetime.strptime(f"{row_id} {date.year}", "%B %Y")
        
        if row_id.month == date.month and row_id.year == date.year:
            log.info(f"{header[1].value.strip()}: {row[1].value:,}")
            
            for i in range(2, 6):
                percentage = row[i].value * 100
                log.info(f"{header[i].value.strip()}: {percentage:.2f}%")
                
            break
    else:
        # if relevant date is not found, log that the data is missing
        log.warning(f"Entry for {date.month} {date.year} not found in {wb.sheetnames[0]}")


def rank_score(score: int, threshold: int) -> str:
    """return a ranking of "good" or "bad" depending on whether score meets threshold

    Args:
        score (int): value to be ranked
        threshold (int): value that score must be equal or greater to to be "good"

    Returns:
        str: "good" or "bad" depending on whether score meets threshold
    """    """  """
    return "good" if score >= threshold else "bad"


def log_tab_two(wb: Workbook, date: datetime) -> None:
    """logs data from second tab of report for given date

    Args:
        wb (Workbook): excel data
        date (datetime): date from file name
    """    
    # NOTE: There is an issue with Pylance and openpyxl. This does not affect execution
    ws: Worksheet = wb[wb.sheetnames[1]]
    
    # iterate through columns containing relevant data
    for col in ws.iter_cols(min_col=2):
        header: datetime = col[0].value
        if header is None:
            # end loop when first empty header is reached
            break
        elif isinstance(header, str):
            # if header is text, convert to datetime object with missing data replaced with data from filename
            header = datetime.strptime(f"{header} {date.year}", "%B %Y")

        if header.month == date.month and header.year == date.year:
            log.info(f"Promoters: {rank_score(col[3].value, 200)}")
            log.info(f"Passives: {rank_score(col[5].value, 100)}")
            log.info(f"Passives: {rank_score(col[7].value, 100)}")
            break
    else:
        # if relevant date is not found, log that the data is missing
        log.warning(f"Entry for {date.month} {date.year} not found in {wb.sheetnames[1]}")


def main() -> None:
    path = WORKING_PATH / DATA_DIR
    errors = False # error flag
    no_files = True # no file flag is True by default
    try:
        for file in path.iterdir():
            if is_valid_filename(file.name):
                print(f"Attempting to log '{file.name}'")
                try:
                    wb = load_workbook(str(file), read_only=False)
                    date = get_date_from_filename(file.name)
                    log.info(date)
                    log_tab_one(wb, date)
                    log_tab_two(wb, date)
                    no_files = False # as soon as file is successfully parsed, set no_file to False
                except Exception as e:
                    # if error occurs while trying to read a specific file, print error to terminal and set errors flag
                    log.warning(e)
                    errors = True
    except FileNotFoundError:
        # catch missing directory error
        error = f"error: cannot read data because directory '{DATA_DIR}' is missing"
        print(error)
        log.error(error)
    except Exception as e:
        # catch any critical errors beyond a missing directory
        print(e)
        log.error(e)
    else:
        if errors:
            status = "LOGGING COMPLETE. SOME FILES COULD NOT BE READ."
        elif no_files:
            status = "LOGGING COMPLETE. NO VALID FILES FOUND."
        else:
            status = "LOGGING COMPLETE."
        print(status)
        log.info(status)

if __name__ == '__main__':
    main()